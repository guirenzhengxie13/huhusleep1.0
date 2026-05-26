from __future__ import annotations

import csv
import io
import json
import re
import sys
from copy import copy
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


BASE_DIR = Path(__file__).resolve().parent
START_DAY = "0515"
INPUT_DIR = BASE_DIR / "输入"
REPORT_DIR = BASE_DIR / "report"
REAL_DIR = BASE_DIR / "real"
RAW_DIR = BASE_DIR / "原始数据"

CITIES = ["姜堰", "合肥", "南京"]

ROW_LABELS = ["空挂床", "正常", "存疑"]
COL_LABELS = ["空挂床", "正常", "存疑"]
CITY_START_COL = {"姜堰": 3, "合肥": 7, "南京": 11}
IDENTITY_RESULT_MAP = {
    "0": "建档",
    "1": "正常",
    "2": "存疑",
    "3": "空挂床",
}
SUPPLEMENTAL_NANJING_ROSTER = {
    "10334b827e3641cd": ("管德莲", "中度失能"),
    "10334b82c88d0130": ("边秋凤", "重度失能"),
    "10334b82bbcac36a": ("李福燮", "重度失能"),
    "10334b82c88d024a": ("王培培", "中度失能"),
    "10334b82c194a349": ("夏京生", "重度失能"),
    "10334b82bbcac122": ("黄永霞", "重度失能"),
    "10334b82c88d036f": ("杨茂梅", "中度失能"),
    "10334b82f89ec3d1": ("华宁", "重度失能"),
    "10334b82e759c2b4": ("徐亚珍", "重度失能"),
    "10334b82c194a0c4": ("张永健", "重度失能"),
}

YELLOW_FILL = PatternFill("solid", fgColor="FFFFFF00")
RED_FILL = PatternFill("solid", fgColor="FFFF0000")
NO_FILL = PatternFill(fill_type=None)
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
GROUP_FILL = PatternFill("solid", fgColor="F3F7FB")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


@dataclass
class Device:
    name: str
    device_id: str
    raw: dict[str, object] = field(default_factory=dict)


@dataclass
class ReportDay:
    city: str
    day: str
    total_count: int = 0
    build_count: int = 0
    invalid_count: int = 0
    empty_count: int = 0
    normal_count: int = 0
    suspect_count: int = 0
    build_devices: list[Device] = field(default_factory=list)
    invalid_devices: list[Device] = field(default_factory=list)
    empty_devices: list[Device] = field(default_factory=list)
    normal_devices: list[Device] = field(default_factory=list)
    suspect_devices: list[Device] = field(default_factory=list)


@dataclass
class TruthRecord:
    city: str
    day: str
    name: str
    device_id: str
    truth: str
    status_text: str
    note: str


@dataclass
class ErrorDetail:
    day: str
    city: str
    truth: str
    parsed: str
    device_id: str
    name: str
    raw: dict[str, object] = field(default_factory=dict)


@dataclass
class ZoneDifference:
    day: str
    city: str
    issue: str
    device_id: str
    name: str


@dataclass
class EvalResult:
    matrix: dict[str, dict[str, int]]
    notes: dict[str, list[str]]
    errors: list[ErrorDetail] = field(default_factory=list)
    missing: bool = False


def norm_id(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    match = re.search(r"([0-9a-f]{16})$", text)
    return match.group(1) if match else text


def validate_day(day: str) -> str:
    if not re.fullmatch(r"\d{4}", day):
        raise ValueError("日期参数必须是四位数字，例如 0515")
    month = int(day[:2])
    day_num = int(day[2:])
    if month < 1 or month > 12 or day_num < 1 or day_num > 31:
        raise ValueError("日期参数格式不合法，例如 0515")
    return day


def configured_start_day() -> str:
    return validate_day(sys.argv[1]) if len(sys.argv) > 1 else validate_day(START_DAY)


def day_sort_key(day: str) -> int:
    return int(day)


def json_dir(start_day: str) -> Path:
    return BASE_DIR / "json" / f"{start_day}_以后"


def output_file(target_days: list[str]) -> Path:
    if not target_days:
        return BASE_DIR / f"{configured_start_day()}以后姜堰合肥南京身份识别及失能等级评估.xlsx"
    return BASE_DIR / f"{target_days[0]}~{target_days[-1]}姜堰合肥南京身份识别及失能等级评估.xlsx"


def parse_count(body: str, item_no: str, label: str) -> int:
    pattern = rf"^{re.escape(item_no)}、共(\d+)台设备{re.escape(label)}"
    match = re.search(pattern, body, flags=re.M)
    return int(match.group(1)) if match else 0


def parse_devices_after_item(body: str, item_no: str) -> list[Device]:
    lines = body.splitlines()
    devices: list[Device] = []
    collecting = False
    for line in lines:
        stripped = line.strip()
        if re.match(r"^\d+(?:\.\d+)?、", stripped):
            collecting = stripped.startswith(f"{item_no}、")
            continue
        if not collecting or not stripped:
            continue
        match = re.search(r"(.+?)([0-9a-fA-F]{16})$", stripped)
        if match:
            devices.append(Device(match.group(1).strip(), match.group(2).lower()))
    return devices


def parse_report_file(path: Path, target_days: list[str]) -> list[ReportDay]:
    text = path.read_text(encoding="utf-8")
    reports: list[ReportDay] = []
    pattern = re.compile(
        r"=+\s*\n"
        r"202605(?P<day>\d{2})(?P<city>姜堰|合肥|南京).*?：\s*\n"
        r"(?P<body>.*?)(?=\*{10,}|\Z)",
        flags=re.S,
    )
    for match in pattern.finditer(text):
        day = day_from_generated_yyyymmdd(f"202605{match.group('day')}")
        city = match.group("city")
        body = match.group("body")
        if target_days and day not in target_days:
            continue
        reports.append(
            ReportDay(
                city=city,
                day=day,
                total_count=parse_count(body, "1", "生成报告"),
                invalid_count=parse_count(body, "2", "的数据无效"),
                empty_count=parse_count(body, "3", "空挂床"),
                normal_count=parse_count(body, "4", "身份识别结果为正常"),
                suspect_count=parse_count(body, "5", "身份识别结果为存疑"),
                invalid_devices=parse_devices_after_item(body, "2"),
                empty_devices=parse_devices_after_item(body, "3"),
                suspect_devices=parse_devices_after_item(body, "5"),
            )
        )
    return reports


def parse_reports() -> dict[tuple[str, str], ReportDay]:
    result: dict[tuple[str, str], ReportDay] = {}
    for path in sorted(REPORT_DIR.rglob("*.txt")):
        for report in parse_report_file(path, []):
            result[(report.city, report.day)] = report
    return result


def parse_zone_difference_file(path: Path, target_days: list[str]) -> list[ZoneDifference]:
    text = path.read_text(encoding="utf-8")
    results: list[ZoneDifference] = []
    pattern = re.compile(
        r"\*+\s*\n"
        r"202605(?P<day>\d{2})(?P<city>姜堰|合肥|南京).*?：\s*\n"
        r"(?P<body>.*?)(?=\n=+|\Z)",
        flags=re.S,
    )
    for match in pattern.finditer(text):
        day = day_from_generated_yyyymmdd(f"202605{match.group('day')}")
        if target_days and day not in target_days:
            continue
        city = match.group("city")
        body = match.group("body")
        for device in parse_devices_after_item(body, "5"):
            results.append(
                ZoneDifference(
                    day=day,
                    city=city,
                    issue="在床分区长不一致",
                    device_id=device.device_id,
                    name=device.name,
                )
            )
    return results


def parse_zone_differences(target_days: list[str]) -> list[ZoneDifference]:
    differences: list[ZoneDifference] = []
    for path in sorted(REPORT_DIR.rglob("*.txt")):
        differences.extend(parse_zone_difference_file(path, target_days))
    return differences


def raw_data_dir() -> Path:
    if INPUT_DIR.exists():
        return INPUT_DIR
    if RAW_DIR.exists():
        return RAW_DIR
    candidates = [path for path in BASE_DIR.iterdir() if path.is_dir() and "数据" in path.name]
    if not candidates:
        raise FileNotFoundError("未找到原始数据文件夹")
    return candidates[0]


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    text = "\n".join(line for line in path.read_text(encoding="utf-8-sig").splitlines() if line.strip())
    if not text:
        return []
    return list(csv.DictReader(io.StringIO(text)))


def day_from_time(value: object) -> str:
    text = "" if value is None else str(value).strip()
    try:
        log_time = datetime.strptime(text[:19], "%Y-%m-%d %H:%M:%S")
    except ValueError:
        return ""
    actual_day = log_time - timedelta(days=1)
    return actual_day.strftime("%m%d")


def day_from_generated_yyyymmdd(value: str) -> str:
    try:
        generated_day = datetime.strptime(value, "%Y%m%d")
    except ValueError:
        return ""
    actual_day = generated_day - timedelta(days=1)
    return actual_day.strftime("%m%d")


def seconds_to_hm(value: object) -> str:
    text = "" if value is None else str(value).strip()
    if text == "":
        return ""
    try:
        total_seconds = int(float(text))
    except ValueError:
        return text
    total_minutes = max(0, total_seconds) // 60
    hours, minutes = divmod(total_minutes, 60)
    return f"{hours}h{minutes}min"


def build_identity_json_files(start_day: str) -> list[str]:
    output_dir = json_dir(start_day)
    output_dir.mkdir(parents=True, exist_ok=True)
    for old_json in output_dir.glob("identity_*.json"):
        old_json.unlink()

    grouped: dict[str, list[dict[str, object]]] = {}
    for path in sorted(raw_data_dir().glob("*.csv")):
        for row in read_csv_rows(path):
            day = day_from_time(row.get("time"))
            device_id = norm_id(row.get("device_id"))
            identity_result = str(row.get("identity_result", "")).strip()
            if not day or day_sort_key(day) < day_sort_key(start_day) or not device_id:
                continue
            grouped.setdefault(day, []).append(
                {
                    "date": day,
                    "source_file": path.name,
                    "device_id": device_id,
                    "time": row.get("time", ""),
                    "identity_result": int(identity_result) if identity_result.isdigit() else None,
                    "identity_label": IDENTITY_RESULT_MAP.get(identity_result, "未知"),
                    "identity_score": row.get("identity_score", ""),
                    "empty_bed_result": row.get("empty_bed_result", ""),
                    "onlinetimes": row.get("onlinetimes", ""),
                    "inbedtimes_day": row.get("inbedtimes_day", ""),
                    "inbedtimes_night": row.get("inbedtimes_night", ""),
                    "sleeptimes_night": row.get("sleeptimes_night", ""),
                    "base_resp_rate": row.get("base_resp_rate", ""),
                    "base_heart_rate": row.get("base_heart_rate", ""),
                    "base_behavior": row.get("base_behavior", ""),
                    "base_update_time": row.get("base_update_time", ""),
                    "resp_rate": row.get("resp_rate", ""),
                    "heart_rate": row.get("heart_rate", ""),
                    "behavior": row.get("behavior", ""),
                    "cur_day_timestamps": row.get("cur_day_timestamps", ""),
                }
            )

    for day, records in grouped.items():
        payload = {
            "date": day,
            "start_day": start_day,
            "identity_result_map": IDENTITY_RESULT_MAP,
            "records": records,
        }
        output = output_dir / f"identity_{day}.json"
        output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return sorted(grouped, key=day_sort_key)


def load_identity_json_records(input_dir: Path) -> dict[str, list[dict[str, object]]]:
    records_by_day: dict[str, list[dict[str, object]]] = {}
    for path in sorted(input_dir.glob("identity_*.json")):
        payload = json.loads(path.read_text(encoding="utf-8"))
        day = str(payload.get("date") or path.stem.replace("identity_", ""))
        records_by_day[day] = list(payload.get("records", []))
    return records_by_day


def reports_from_identity_jsons(
    truths: dict[tuple[str, str], dict[str, TruthRecord]],
    records_by_day: dict[str, list[dict[str, object]]],
) -> dict[tuple[str, str], ReportDay]:
    reports: dict[tuple[str, str], ReportDay] = {}

    for day in sorted(records_by_day, key=day_sort_key):
        raw_records = records_by_day.get(day, [])
        for city in CITIES:
            truth_map = truths.get((city, day), {})
            if not truth_map:
                continue

            matched: dict[str, dict[str, object]] = {}
            for record in raw_records:
                device_id = norm_id(record.get("device_id"))
                truth = truth_map.get(device_id)
                if truth:
                    matched[truth.device_id] = record

            if not matched:
                continue

            report = ReportDay(city=city, day=day, total_count=len(matched))
            for device_id, record in matched.items():
                truth = truth_map.get(device_id)
                name = truth.name if truth else ""
                device = Device(name=name, device_id=device_id, raw=record)
                label = str(record.get("identity_label") or "")
                if label == "建档":
                    report.build_count += 1
                    report.build_devices.append(device)
                elif label == "正常":
                    report.normal_count += 1
                    report.normal_devices.append(device)
                elif label == "存疑":
                    report.suspect_count += 1
                    report.suspect_devices.append(device)
                elif label == "空挂床":
                    report.empty_count += 1
                    report.empty_devices.append(device)

            reports[(city, day)] = report
    return reports


def infer_city_from_filename(path: Path) -> str | None:
    name = path.name
    for city in CITIES:
        if city in name:
            return city
    return None


def header_index(headers: list[object], keyword: str) -> int | None:
    for idx, header in enumerate(headers, start=1):
        if header and keyword in str(header):
            return idx
    return None


def truth_from_row(status: object, row_values: list[object]) -> str:
    status_text = "" if status is None else str(status).strip()
    explicit_suspect = any(value is not None and "存疑" in str(value) for value in row_values)

    if any(key in status_text for key in ["空床", "离院", "去世", "离线"]):
        return "空挂床"
    if explicit_suspect:
        return "存疑"
    return "正常"


def load_truths(target_days: list[str]) -> dict[tuple[str, str], dict[str, TruthRecord]]:
    truths: dict[tuple[str, str], dict[str, TruthRecord]] = {}
    for path in sorted(REAL_DIR.glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        city = infer_city_from_filename(path)
        if city is None:
            continue

        workbook = load_workbook(path, data_only=True)
        sheet = workbook.worksheets[0]
        headers = [cell.value for cell in sheet[1]]
        device_col = header_index(headers, "设备编号")
        collect_col = header_index(headers, "采集设备编号")
        name_col = header_index(headers, "老人姓名")
        change_col = header_index(headers, "床位调换")
        if not device_col or not name_col:
            continue

        day_cols: dict[str, int] = {}
        for idx, header in enumerate(headers, start=1):
            text = "" if header is None else str(header)
            for day in target_days:
                if day in text:
                    day_cols[day] = idx

        for row in range(2, sheet.max_row + 1):
            primary_id = norm_id(sheet.cell(row, device_col).value)
            if not primary_id:
                continue
            aliases = {primary_id}
            if collect_col:
                collect_id = norm_id(sheet.cell(row, collect_col).value)
                if collect_id:
                    aliases.add(collect_id)

            name = str(sheet.cell(row, name_col).value or "").strip()
            row_values = [sheet.cell(row, col).value for col in range(1, sheet.max_column + 1)]
            change_note = sheet.cell(row, change_col).value if change_col else None
            for day, col in day_cols.items():
                status = sheet.cell(row, col).value
                truth = truth_from_row(status, row_values)
                record = TruthRecord(
                    city=city,
                    day=day,
                    name=name,
                    device_id=primary_id,
                    truth=truth,
                    status_text=str(status or "").strip(),
                    note=str(change_note or "").strip(),
                )
                bucket = truths.setdefault((city, day), {})
                for alias in aliases:
                    bucket[alias] = record

    for day in target_days:
        bucket = truths.setdefault(("南京", day), {})
        for device_id, (name, disability_level) in SUPPLEMENTAL_NANJING_ROSTER.items():
            bucket[device_id] = TruthRecord(
                city="南京",
                day=day,
                name=name,
                device_id=device_id,
                truth="正常",
                status_text="在院",
                note=disability_level,
            )
    return truths


def add_hefei_truths_from_jsons(
    truths: dict[tuple[str, str], dict[str, TruthRecord]],
    records_by_day: dict[str, list[dict[str, object]]],
) -> None:
    for day, records in records_by_day.items():
        known_ids = {
            alias
            for city in CITIES
            for alias in truths.get((city, day), {})
        }
        bucket = truths.setdefault(("合肥", day), {})
        for record in records:
            device_id = norm_id(record.get("device_id"))
            if not device_id or device_id in known_ids:
                continue
            bucket[device_id] = TruthRecord(
                city="合肥",
                day=day,
                name="",
                device_id=device_id,
                truth="正常",
                status_text="在院",
                note="补充映射：非姜堰、非南京设备归为合肥",
            )


def empty_matrix() -> dict[str, dict[str, int]]:
    return {row: {col: 0 for col in COL_LABELS} for row in ROW_LABELS}


def unique_truth_records(truth_map: dict[str, TruthRecord]) -> list[TruthRecord]:
    seen: set[str] = set()
    records: list[TruthRecord] = []
    for record in truth_map.values():
        if record.device_id in seen:
            continue
        seen.add(record.device_id)
        records.append(record)
    return records


def device_label(device: Device, record: TruthRecord | None = None) -> str:
    if record:
        name = record.name or device.name
        return f"{name}{record.device_id}"
    return f"{device.name}{device.device_id}"


def evaluate_report(report: ReportDay | None, truth_map: dict[str, TruthRecord] | None) -> EvalResult:
    matrix = empty_matrix()
    notes = {row: [] for row in ROW_LABELS}
    errors: list[ErrorDetail] = []
    if report is None:
        return EvalResult(matrix=matrix, notes=notes, missing=True)

    if not truth_map:
        matrix["空挂床"]["空挂床"] = report.empty_count
        matrix["正常"]["正常"] = report.normal_count
        matrix["存疑"]["存疑"] = report.suspect_count
        return EvalResult(matrix=matrix, notes=notes)

    def add_known(device: Device, parsed: str) -> None:
        record = truth_map.get(norm_id(device.device_id))
        truth = record.truth if record else "正常"
        if parsed in COL_LABELS:
            matrix[truth][parsed] += 1
        if truth != parsed:
            notes[truth].append(f"{truth}识别为{parsed}：{device_label(device, record)}")
            errors.append(
                ErrorDetail(
                    day=report.day,
                    city=report.city,
                    truth=truth,
                    parsed=parsed,
                    device_id=record.device_id if record else norm_id(device.device_id),
                    name=record.name if record and record.name else device.name,
                    raw=device.raw,
                )
            )

    for device in report.empty_devices:
        add_known(device, "空挂床")

    for device in report.normal_devices:
        add_known(device, "正常")

    for device in report.suspect_devices:
        add_known(device, "存疑")

    # identity_result=0 is the initial building-profile state, not an identity
    # recognition result. It is intentionally excluded from both the matrix and
    # the error detail table.
    return EvalResult(matrix=matrix, notes=notes, errors=errors)


def find_template() -> Path:
    candidates = sorted(BASE_DIR.glob("0505~0509*.xlsx"))
    if not candidates:
        raise FileNotFoundError("未找到 0505~0509 模板 xlsx")
    return candidates[0]


def set_cell(cell, value, source_cell=None) -> None:
    cell.value = value
    if source_cell is not None:
        cell.font = copy(source_cell.font)
        cell.alignment = copy(source_cell.alignment)
        cell.border = copy(source_cell.border)
        cell.fill = copy(source_cell.fill)
        cell.number_format = source_cell.number_format


def prepare_identity_sheet(sheet, target_days: list[str]) -> None:
    required_days = max(len(target_days), 5)
    required_last_row = 3 + required_days * 3
    if required_days > 5 and sheet.max_row < required_last_row:
        sheet.insert_rows(19, required_last_row - 18)

    for row in sheet.iter_rows(min_row=1, max_row=required_last_row, min_col=1, max_col=14):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            cell.value = None
            if cell.row >= 4 and cell.column >= 3:
                cell.fill = copy(NO_FILL)

    title = f"{target_days[0]}~{target_days[-1]}身份识别" if target_days else "身份识别"
    sheet["A1"] = title
    sheet["C1"] = "姜堰"
    sheet["G1"] = "合肥"
    sheet["K1"] = "南京"
    for col in [3, 7, 11]:
        sheet.cell(2, col).value = "有效数据报告"
        sheet.cell(3, col).value = "识别为空挂床"
        sheet.cell(3, col + 1).value = "识别为正常"
        sheet.cell(3, col + 2).value = "识别为存疑"
        sheet.cell(3, col + 3).value = "备注"

    existing_merges = {str(rng) for rng in sheet.merged_cells.ranges}
    for idx, day in enumerate(target_days):
        row = 4 + idx * 3
        date_merge = f"A{row}:A{row + 2}"
        if date_merge not in existing_merges:
            sheet.merge_cells(date_merge)
        sheet.cell(row, 1).value = day
        sheet.cell(row, 2).value = "真实空挂床"
        sheet.cell(row + 1, 2).value = "真实正常"
        sheet.cell(row + 2, 2).value = "真实存疑"

    for row in sheet.iter_rows(min_row=1, max_row=required_last_row, min_col=1, max_col=14):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if cell.row <= 3 or cell.column <= 2:
                cell.font = Font(name="Calibri", bold=True)


def write_identity_results(sheet, reports, truths, target_days: list[str]) -> list[ErrorDetail]:
    all_errors: list[ErrorDetail] = []
    for day_idx, day in enumerate(target_days):
        base_row = 4 + day_idx * 3
        for city in CITIES:
            start_col = CITY_START_COL[city]
            report = reports.get((city, day))
            truth_map = truths.get((city, day))
            result = evaluate_report(report, truth_map)
            all_errors.extend(result.errors)

            for row_offset, truth in enumerate(ROW_LABELS):
                row = base_row + row_offset
                for col_offset, parsed in enumerate(COL_LABELS):
                    cell = sheet.cell(row, start_col + col_offset)
                    if result.missing:
                        cell.value = None
                        cell.fill = copy(NO_FILL)
                    else:
                        cell.value = result.matrix[truth][parsed]
                        cell.fill = copy(NO_FILL)

                remark = sheet.cell(row, start_col + 3)
                remark.value = None
                remark.fill = copy(NO_FILL)

            normal_as_suspect = sheet.cell(base_row + 1, start_col + 2)
            suspect_as_normal = sheet.cell(base_row + 2, start_col + 1)
            if normal_as_suspect.value:
                normal_as_suspect.fill = copy(YELLOW_FILL)
            if suspect_as_normal.value:
                suspect_as_normal.fill = copy(RED_FILL)
    return all_errors


def sorted_error_details(errors: list[ErrorDetail]) -> list[ErrorDetail]:
    city_order = {city: idx for idx, city in enumerate(CITIES)}
    label_order = {label: idx for idx, label in enumerate(ROW_LABELS)}
    return sorted(
        errors,
        key=lambda item: (
            item.day,
            city_order.get(item.city, 99),
            label_order.get(item.truth, 99),
            label_order.get(item.parsed, 99),
            item.device_id,
        ),
    )


def merge_same_values(sheet, start_row: int, end_row: int, col: int, key_cols: list[int]) -> None:
    row = start_row
    while row <= end_row:
        next_row = row + 1
        while next_row <= end_row and all(
            sheet.cell(next_row, key_col).value == sheet.cell(row, key_col).value for key_col in key_cols
        ):
            next_row += 1
        if next_row - row > 1:
            sheet.merge_cells(start_row=row, start_column=col, end_row=next_row - 1, end_column=col)
        row = next_row


def create_error_detail_sheet(workbook, errors: list[ErrorDetail]) -> None:
    if "识别错误明细" in workbook.sheetnames:
        workbook.remove(workbook["识别错误明细"])
    sheet = workbook.create_sheet("识别错误明细")

    sorted_errors = sorted_error_details(errors)
    sheet.merge_cells("A1:Q1")
    sheet["A1"] = "识别错误明细"
    sheet["A1"].font = Font(name="Calibri", bold=True, size=15)
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet["A1"].fill = copy(HEADER_FILL)
    sheet.row_dimensions[1].height = 30

    headers = [
        "日期",
        "城市",
        "真实情况",
        "识别结果",
        "设备号",
        "老人姓名",
        "身份识别得分",
        "日间在床时长",
        "夜间在床时长",
        "夜间睡眠时长",
        "基础呼吸率",
        "基础心率",
        "基础行为",
        "基线更新时间",
        "呼吸率数组",
        "心率数组",
        "行为数组",
    ]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(2, col)
        cell.value = header
        cell.font = Font(name="Calibri", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = copy(HEADER_FILL)
        cell.border = copy(THIN_BORDER)
    sheet.row_dimensions[2].height = 24

    if not sorted_errors:
        sheet.merge_cells("A3:Q3")
        sheet["A3"] = "无识别错误"
        sheet["A3"].alignment = Alignment(horizontal="center", vertical="center")
        sheet["A3"].border = copy(THIN_BORDER)
        return

    for row_offset, error in enumerate(sorted_errors, start=3):
        row = row_offset
        values = [
            error.day,
            error.city,
            f"真实{error.truth}",
            f"识别为{error.parsed}",
            error.device_id,
            error.name,
            error.raw.get("identity_score", ""),
            seconds_to_hm(error.raw.get("inbedtimes_day")),
            seconds_to_hm(error.raw.get("inbedtimes_night")),
            seconds_to_hm(error.raw.get("sleeptimes_night")),
            error.raw.get("base_resp_rate", ""),
            error.raw.get("base_heart_rate", ""),
            error.raw.get("base_behavior", ""),
            error.raw.get("base_update_time", ""),
            error.raw.get("resp_rate", ""),
            error.raw.get("heart_rate", ""),
            error.raw.get("behavior", ""),
        ]
        for col, value in enumerate(values, start=1):
            cell = sheet.cell(row, col)
            cell.value = value
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = copy(THIN_BORDER)
            if col <= 4:
                cell.fill = copy(GROUP_FILL)
            elif 8 <= col <= 10:
                cell.fill = PatternFill("solid", fgColor="D9EAF7")
            elif col >= 11:
                cell.fill = PatternFill("solid", fgColor="C6E97A")
        sheet.row_dimensions[row].height = 28

    end_row = 2 + len(sorted_errors)
    merge_same_values(sheet, 3, end_row, 4, [1, 2, 3, 4])
    merge_same_values(sheet, 3, end_row, 3, [1, 2, 3])
    merge_same_values(sheet, 3, end_row, 2, [1, 2])
    merge_same_values(sheet, 3, end_row, 1, [1])

    sheet.column_dimensions["A"].width = 11
    sheet.column_dimensions["B"].width = 12
    sheet.column_dimensions["C"].width = 16
    sheet.column_dimensions["D"].width = 16
    sheet.column_dimensions["E"].width = 22
    sheet.column_dimensions["F"].width = 18
    sheet.column_dimensions["G"].width = 14
    sheet.column_dimensions["H"].width = 14
    sheet.column_dimensions["I"].width = 14
    sheet.column_dimensions["J"].width = 14
    sheet.column_dimensions["K"].width = 12
    sheet.column_dimensions["L"].width = 12
    sheet.column_dimensions["M"].width = 18
    sheet.column_dimensions["N"].width = 20
    sheet.column_dimensions["O"].width = 36
    sheet.column_dimensions["P"].width = 36
    sheet.column_dimensions["Q"].width = 36
    sheet.freeze_panes = "A3"


def create_zone_difference_sheet(workbook, differences: list[ZoneDifference]) -> None:
    if "在床分区长不一致" in workbook.sheetnames:
        workbook.remove(workbook["在床分区长不一致"])
    sheet = workbook.create_sheet("在床分区长不一致")

    city_order = {city: idx for idx, city in enumerate(CITIES)}
    sorted_differences = sorted(
        differences,
        key=lambda item: (item.day, city_order.get(item.city, 99), item.issue, item.device_id),
    )

    sheet.merge_cells("A1:E1")
    sheet["A1"] = "在床分区长不一致"
    sheet["A1"].font = Font(name="Calibri", bold=True, size=15)
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet["A1"].fill = copy(HEADER_FILL)
    sheet.row_dimensions[1].height = 30

    headers = ["日期", "城市", "问题类型", "设备号", "老人姓名"]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(2, col)
        cell.value = header
        cell.font = Font(name="Calibri", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = copy(HEADER_FILL)
        cell.border = copy(THIN_BORDER)
    sheet.row_dimensions[2].height = 24

    if not sorted_differences:
        sheet.merge_cells("A3:E3")
        sheet["A3"] = "无在床分区长不一致"
        sheet["A3"].alignment = Alignment(horizontal="center", vertical="center")
        sheet["A3"].border = copy(THIN_BORDER)
        return

    for row, item in enumerate(sorted_differences, start=3):
        values = [item.day, item.city, item.issue, item.device_id, item.name]
        for col, value in enumerate(values, start=1):
            cell = sheet.cell(row, col)
            cell.value = value
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = copy(THIN_BORDER)
            if col <= 3:
                cell.fill = copy(GROUP_FILL)
        sheet.row_dimensions[row].height = 28

    end_row = 2 + len(sorted_differences)
    merge_same_values(sheet, 3, end_row, 3, [1, 2, 3])
    merge_same_values(sheet, 3, end_row, 2, [1, 2])
    merge_same_values(sheet, 3, end_row, 1, [1])

    sheet.column_dimensions["A"].width = 11
    sheet.column_dimensions["B"].width = 12
    sheet.column_dimensions["C"].width = 22
    sheet.column_dimensions["D"].width = 22
    sheet.column_dimensions["E"].width = 18
    sheet.freeze_panes = "A3"


def remove_extra_sheets(workbook) -> None:
    for sheet in list(workbook.worksheets[1:]):
        workbook.remove(sheet)


def save_output(workbook, output_path: Path) -> Path:
    try:
        workbook.save(output_path)
        return output_path
    except PermissionError:
        fallback = output_path.with_name(f"{output_path.stem}_new{output_path.suffix}")
        workbook.save(fallback)
        return fallback


def main() -> None:
    start_day = configured_start_day()
    target_days = build_identity_json_files(start_day)
    json_path = json_dir(start_day)
    records_by_day = load_identity_json_records(json_path)
    if not target_days:
        target_days = sorted(records_by_day, key=day_sort_key)

    truths = load_truths(target_days)
    add_hefei_truths_from_jsons(truths, records_by_day)
    reports = reports_from_identity_jsons(truths, records_by_day)

    workbook = load_workbook(find_template())
    identity_sheet = workbook.worksheets[0]
    prepare_identity_sheet(identity_sheet, target_days)
    errors = write_identity_results(identity_sheet, reports, truths, target_days)
    zone_differences = parse_zone_differences(target_days)
    remove_extra_sheets(workbook)
    create_error_detail_sheet(workbook, errors)
    create_zone_difference_sheet(workbook, zone_differences)

    output = save_output(workbook, output_file(target_days))
    print(f"生成完成：{output}")
    print(f"JSON目录：{json_path}")
    missing = [(city, day) for day in target_days for city in CITIES if (city, day) not in reports]
    if missing:
        print("缺少解析数据：")
        for city, day in missing:
            print(f"  {day} {city}")


if __name__ == "__main__":
    main()
