import argparse
import csv
import json
import logging
import os
import re
import shutil
import sys
from collections import defaultdict
from datetime import datetime

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from utils import build_device_to_location_from_roster, get_device_roster


DEFAULT_IMPORT_DIR = r"C:\Users\Lenovo\Downloads"
DEFAULT_OUTPUT_ROOT = r"C:\Users\Lenovo\Desktop\data\测试跟踪"
CONFIG_FILE_PATH = os.path.join(PROJECT_ROOT, "config.json")
TRACKED_LOCATION_CODES = {"hf", "jy", "nj"}

DATA_TYPE_DEVICE_STATUS = "device_status"
DATA_TYPE_IDENTITY_2D43 = "identity_2d43"

DATA_TYPE_DIRS = {
    DATA_TYPE_DEVICE_STATUS: "device_status",
    DATA_TYPE_IDENTITY_2D43: "identity_2d43",
}

DATA_TYPE_FILENAMES = {
    DATA_TYPE_DEVICE_STATUS: "设备情况表",
    DATA_TYPE_IDENTITY_2D43: "2.d.43",
}

MANUAL_REPORT_FIELDS = [
    "院区",
    "姓名",
    "device_id",
    "日期",
    "time",
    "cur_day_timestamps",
    "state_day",
    "补充值",
    "备注",
]


def load_config_data(config_file_path=CONFIG_FILE_PATH):
    with open(config_file_path, "r", encoding="utf-8") as f:
        return json.load(f)


def _normalize_cell(value):
    return str(value).strip().lstrip("\ufeff") if value is not None else ""


def _read_csv_rows(csv_path):
    last_error = None
    for encoding in ("utf-8-sig", "gbk"):
        try:
            with open(csv_path, "r", encoding=encoding, errors="strict", newline="") as f:
                rows = [
                    [_normalize_cell(cell) for cell in row]
                    for row in csv.reader(f)
                    if any(_normalize_cell(cell) for cell in row)
                ]
            return rows
        except UnicodeDecodeError as e:
            last_error = e

    raise last_error


def _records_from_rows(rows):
    if not rows:
        return [], []
    header = rows[0]
    records = []
    for row in rows[1:]:
        record = {}
        for index, column in enumerate(header):
            record[column] = row[index] if index < len(row) else ""
        records.append(record)
    return header, records


def _detect_data_type(header):
    header_set = set(header)
    if {"device_id", "state_day"}.issubset(header_set):
        return DATA_TYPE_IDENTITY_2D43
    if "设备编号" in header_set and any("身份识别日报" in col or "设备异常重启" in col for col in header):
        return DATA_TYPE_DEVICE_STATUS
    return None


def build_device_location_index(config_data, project_root):
    roster_path = os.path.join(project_root, "assets", "full_device_roster.csv")
    return {
        device_id: location_code
        for device_id, location_code in build_device_to_location_from_roster(config_data, roster_path).items()
        if location_code in TRACKED_LOCATION_CODES
    }


def build_device_name_index(project_root):
    roster_path = os.path.join(project_root, "assets", "full_device_roster.csv")
    return {
        row["设备号"]: row["老人姓名"]
        for row in get_device_roster(roster_path)
        if row["设备号"] and row["老人姓名"]
    }


def _device_id_for_record(record, data_type):
    if data_type == DATA_TYPE_DEVICE_STATUS:
        return record.get("设备编号", "").strip()
    return record.get("device_id", "").strip()


def _parse_date_text(value):
    value = (value or "").strip()
    if not value or value == "-1":
        return None

    if re.fullmatch(r"20\d{6}", value):
        return value

    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d %H:%M:%S", "%Y/%m/%d"):
        try:
            return datetime.strptime(value[:19] if "%H" in fmt else value[:10], fmt).strftime("%Y%m%d")
        except ValueError:
            continue
    return None


def _date_from_filename(filename):
    match = re.search(r"(20\d{6})", filename)
    return match.group(1) if match else None


def _date_for_record(record, data_type, source_filename):
    if data_type == DATA_TYPE_IDENTITY_2D43:
        return (
            _parse_date_text(record.get("time"))
            or _parse_date_text(record.get("cur_day_timestamps"))
            or _date_from_filename(source_filename)
        )
    if data_type == DATA_TYPE_DEVICE_STATUS:
        return _parse_date_text(record.get("日期")) or _date_from_filename(source_filename)
    return _date_from_filename(source_filename)


def inspect_test_tracking_csv(csv_path, device_to_location):
    try:
        rows = _read_csv_rows(csv_path)
    except Exception as e:
        return None, f"读取失败: {e}"

    header, records = _records_from_rows(rows)
    data_type = _detect_data_type(header)
    if not data_type:
        return None, "不是测试跟踪设备情况表或 2.D.43 表"

    location_counter = defaultdict(int)
    date_counter = defaultdict(int)
    recognized_records = 0
    for record in records:
        device_id = _device_id_for_record(record, data_type)
        location_code = device_to_location.get(device_id)
        if not location_code:
            continue
        recognized_records += 1
        location_counter[location_code] += 1
        date_key = _date_for_record(record, data_type, os.path.basename(csv_path))
        if date_key:
            date_counter[date_key] += 1

    if not recognized_records:
        return None, "没有匹配到合肥/姜堰/南京测试设备"

    return {
        "source_path": csv_path,
        "header": header,
        "records": records,
        "data_type": data_type,
        "location_counts": dict(location_counter),
        "date_counts": dict(date_counter),
    }, None


def _location_name(config_data, location_code):
    return config_data.get(location_code, {}).get("name", location_code).replace("院区", "")


def _unique_path(path):
    if not os.path.exists(path):
        return path

    root, ext = os.path.splitext(path)
    index = 2
    while True:
        candidate = f"{root}_{index}{ext}"
        if not os.path.exists(candidate):
            return candidate
        index += 1


def _write_csv(path, fieldnames, rows):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _most_common_non_empty(values, default=""):
    counter = defaultdict(int)
    for value in values:
        text = _normalize_cell(value)
        if text:
            counter[text] += 1
    if not counter:
        return default
    return sorted(counter.items(), key=lambda item: item[1], reverse=True)[0][0]


def _write_device_status_xlsx(path, header, rows, device_to_name):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"

    output_columns = [column for column in header if column != "日期"]
    output_header = ["姓名"] + output_columns
    sheet.append(output_header)

    for record in rows:
        device_id = record.get("设备编号", "")
        sheet.append([device_to_name.get(device_id, "")] + [record.get(column, "") for column in output_columns])

    red_bold_font = Font(color="FF0000", bold=True)
    red_font = Font(color="FF0000")
    bold_font = Font(bold=True)
    header_alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    center_alignment = Alignment(horizontal="center", vertical="center")

    for col in range(1, sheet.max_column + 1):
        sheet.cell(row=1, column=col).alignment = header_alignment
        sheet.cell(row=1, column=col).font = bold_font

    last_data_row = sheet.max_row
    for row in range(2, last_data_row + 1):
        for col in range(3, 7):
            sheet.cell(row=row, column=col).alignment = center_alignment

        cell_c = sheet.cell(row=row, column=3)
        if "有" in _normalize_cell(cell_c.value):
            cell_c.font = red_bold_font

        for col in (4, 5, 6):
            cell = sheet.cell(row=row, column=col)
            value = _normalize_cell(cell.value)
            if "无" in value or value in {"None", "nan", ""}:
                cell.value = "无"
                cell.font = red_font

    summary_row = last_data_row + 1
    sheet.cell(row=summary_row, column=1, value="统计：").alignment = center_alignment
    sheet.cell(row=summary_row, column=1).font = bold_font

    sheet.cell(row=summary_row, column=3, value=f'=COUNTIF(C2:C{last_data_row}, "*有*")').alignment = center_alignment
    for col in (4, 5, 6):
        col_letter = sheet.cell(row=1, column=col).column_letter
        sheet.cell(row=summary_row, column=col, value=f'=COUNTIF({col_letter}2:{col_letter}{last_data_row}, "*无*")').alignment = center_alignment
        sheet.cell(row=summary_row, column=col).font = bold_font
    sheet.cell(row=summary_row, column=3).font = bold_font

    for col_cells in sheet.columns:
        sheet.column_dimensions[col_cells[0].column_letter].width = 18
    sheet.column_dimensions["A"].width = 10
    sheet.column_dimensions["C"].width = 20
    sheet.column_dimensions["F"].width = 25

    workbook.save(path)


def _archive_source_file(source_path, output_root, dry_run=False):
    target_dir = os.path.join(output_root, "source_files")
    os.makedirs(target_dir, exist_ok=True)
    target_path = _unique_path(os.path.join(target_dir, os.path.basename(source_path)))
    if not dry_run:
        shutil.move(source_path, target_path)
    return target_path


def _rows_by_location_and_date(info, config_data, device_to_location, device_to_name):
    grouped = defaultdict(list)
    manual_rows = defaultdict(list)
    data_type = info["data_type"]

    for record in info["records"]:
        device_id = _device_id_for_record(record, data_type)
        location_code = device_to_location.get(device_id)
        if not location_code:
            continue

        date_key = _date_for_record(record, data_type, os.path.basename(info["source_path"]))
        if not date_key:
            date_key = datetime.now().strftime("%Y%m%d")

        archive_record = dict(record)
        grouped[(location_code, date_key)].append(archive_record)

        if data_type == DATA_TYPE_IDENTITY_2D43 and record.get("state_day", "").strip() == "-1":
            manual_rows[(location_code, date_key)].append({
                "院区": _location_name(config_data, location_code),
                "姓名": device_to_name.get(device_id, ""),
                "device_id": device_id,
                "日期": date_key,
                "time": record.get("time", ""),
                "cur_day_timestamps": record.get("cur_day_timestamps", ""),
                "state_day": record.get("state_day", ""),
                "补充值": "",
                "备注": "",
            })

    return grouped, manual_rows


def archive_test_tracking_csv(info, config_data, device_to_location, device_to_name, output_root=DEFAULT_OUTPUT_ROOT, dry_run=False):
    archive_paths = []
    manual_report_paths = []
    data_type = info["data_type"]
    grouped, manual_rows = _rows_by_location_and_date(info, config_data, device_to_location, device_to_name)

    for (location_code, date_key), rows in sorted(grouped.items()):
        location_name = _location_name(config_data, location_code)
        data_dir = DATA_TYPE_DIRS[data_type]
        if data_type == DATA_TYPE_IDENTITY_2D43:
            filename = f"{date_key}{location_name}{DATA_TYPE_FILENAMES[data_type]}.csv"
        else:
            version = _most_common_non_empty([row.get("base_version 2.1.3", "") for row in rows], default="unknown")
            filename = f"{date_key}{location_name}测试情况跟踪-{version}.xlsx"
        target_path = _unique_path(os.path.join(output_root, location_name, data_dir, filename))

        if not dry_run:
            if data_type == DATA_TYPE_IDENTITY_2D43:
                _write_csv(target_path, info["header"], rows)
            else:
                _write_device_status_xlsx(target_path, info["header"], rows, device_to_name)
        archive_paths.append(target_path)

    for (location_code, date_key), rows in sorted(manual_rows.items()):
        location_name = _location_name(config_data, location_code)
        filename = f"{date_key}_state_day待补充.csv"
        target_path = _unique_path(os.path.join(output_root, location_name, "manual_reports", filename))
        if not dry_run:
            _write_csv(target_path, MANUAL_REPORT_FIELDS, rows)
        manual_report_paths.append(target_path)

    source_archive_path = _archive_source_file(info["source_path"], output_root, dry_run=dry_run)
    return {
        "source_archive_path": source_archive_path,
        "archive_paths": archive_paths,
        "manual_report_paths": manual_report_paths,
    }


def discover_and_archive_test_tracking(import_dir=DEFAULT_IMPORT_DIR, output_root=DEFAULT_OUTPUT_ROOT, config_file_path=CONFIG_FILE_PATH, dry_run=False):
    project_root = PROJECT_ROOT
    config_data = load_config_data(config_file_path)
    device_to_location = build_device_location_index(config_data, project_root)
    device_to_name = build_device_name_index(project_root)

    results = []
    skipped = []
    for filename in sorted(os.listdir(import_dir)):
        if not filename.lower().endswith(".csv"):
            continue

        csv_path = os.path.join(import_dir, filename)
        info, reason = inspect_test_tracking_csv(csv_path, device_to_location)
        if not info:
            skipped.append({"source_path": csv_path, "reason": reason})
            logging.info("跳过测试跟踪 CSV: %s | %s", filename, reason)
            continue

        result = archive_test_tracking_csv(
            info=info,
            config_data=config_data,
            device_to_location=device_to_location,
            device_to_name=device_to_name,
            output_root=output_root,
            dry_run=dry_run,
        )
        results.append({"source_path": csv_path, "info": info, **result})
        logging.info("已处理测试跟踪 CSV: %s", filename)

    return {"processed": results, "skipped": skipped}


def run(import_dir=DEFAULT_IMPORT_DIR, output_root=DEFAULT_OUTPUT_ROOT, dry_run=False):
    return discover_and_archive_test_tracking(import_dir=import_dir, output_root=output_root, dry_run=dry_run)


def main():
    parser = argparse.ArgumentParser(description="测试跟踪 CSV 识别、归档和 state_day 待补清单生成")
    parser.add_argument("--import-dir", default=DEFAULT_IMPORT_DIR, help="测试 CSV 导入目录，默认读取 Downloads")
    parser.add_argument("--output-root", default=DEFAULT_OUTPUT_ROOT, help="测试跟踪归档根目录")
    parser.add_argument("--dry-run", action="store_true", help="只识别和打印结果，不写文件、不移动源文件")
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s", datefmt="%H:%M:%S")
    result = run(import_dir=args.import_dir, output_root=args.output_root, dry_run=args.dry_run)

    print("processed:", len(result["processed"]))
    for item in result["processed"]:
        info = item["info"]
        print(os.path.basename(item["source_path"]), info["data_type"], info["location_counts"], info["date_counts"])
        for path in item["archive_paths"]:
            print("  archive:", path)
        for path in item["manual_report_paths"]:
            print("  manual:", path)
        print("  source:", item["source_archive_path"])

    print("skipped:", len(result["skipped"]))
    for item in result["skipped"]:
        print("  skip:", os.path.basename(item["source_path"]), item["reason"])


if __name__ == "__main__":
    main()
