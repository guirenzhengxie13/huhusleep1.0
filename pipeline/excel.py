import os
import re
import csv
import json
import openpyxl
import logging
from collections import defaultdict
from openpyxl.styles import Alignment, Border, PatternFill, Side
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from datetime import datetime

from utils import clean_name, get_device_mapping

def generate_diagnostics(sleep_info, accurate_events, alert_dts, ref_year):
    diagnostics = []
    sleep_periods = []
    for period in sleep_info.get('periods', []):
        try:
            sleep_start_dt = datetime.strptime(f"{ref_year}-{period['start']}", "%Y-%m-%d %H:%M:%S")
            sleep_wake_dt = datetime.strptime(f"{ref_year}-{period['end']}", "%Y-%m-%d %H:%M:%S")
            sleep_periods.append((sleep_start_dt, sleep_wake_dt))
        except Exception:
            continue

    matched_alerts = set() 
    for ev in accurate_events:
        ev_time_str = ev['start_time']
        try:
            ev_dt = datetime.strptime(f"{ref_year}-{ev_time_str}", "%Y-%m-%d %H:%M:%S")
        except Exception:
            continue
            
        duration = float(ev.get('duration', 0))
        if not (ev_dt.hour >= 21 or ev_dt.hour < 6):
            diagnostics.append(f"{ev_time_str}离床时间不在离床预警时间范围内，睡眠报告记录离床")
            continue

        found_alert_dt = None
        for adt in alert_dts:
            if adt in matched_alerts: continue
            if 0 <= (adt - ev_dt).total_seconds() <= 480:  
                found_alert_dt = adt
                break

        if found_alert_dt:
            delta_sec = (found_alert_dt - ev_dt).total_seconds()
            diagnostics.append(f"{ev_time_str}离床，{'30秒' if delta_sec <= 90 else '5分钟'}后离床预警，结果一致")
            matched_alerts.add(found_alert_dt)
        else:
            diagnostics.append(f"{ev_time_str}离床{'5分钟内回床没有预警' if duration <= 5 else '，5分钟后超过预警时间范围，没有预警'}")

    for adt in alert_dts:
        if adt not in matched_alerts:
            adt_str = adt.strftime('%Y-%m-%d %H:%M:%S')
            if sleep_periods:
                in_any_sleep_period = any(start_dt <= adt <= end_dt for start_dt, end_dt in sleep_periods)
                diagnostics.append(f"{adt_str}离床预警不在睡眠时间段内，睡眠报告没有离床" if not in_any_sleep_period else f"{adt_str}有离床预警，睡眠报告没有记录")
            else:
                diagnostics.append(f"{adt_str}有离床预警，但睡眠报告未生成")
    return diagnostics


def _append_sleep_record(sleep_data, name, floor, sleep_start, sleep_end, leave_bed):
    item = sleep_data.setdefault(name, {
        'periods': [],
        'leave_bed_times': [],
        'floor': floor,
    })
    if floor and not item.get('floor'):
        item['floor'] = floor
    item['periods'].append({"start": sleep_start, "end": sleep_end})
    if leave_bed:
        item['leave_bed_times'].extend([value for value in leave_bed.split('\n') if value.strip()])


def _finalize_sleep_data(sleep_data):
    for item in sleep_data.values():
        item['periods'].sort(key=lambda value: value['start'])
        item['leave_bed_times'].sort()
        sleep_lines = []
        for period in item['periods']:
            sleep_lines.append(f"入睡 {period['start']}")
            sleep_lines.append(f"清醒 {period['end']}")
        item['入睡清醒'] = '\n'.join(sleep_lines)
        item['离床时间'] = '\n'.join(item['leave_bed_times'])
        item['e_离床时间'] = '\n'.join([f"{value} 离床分钟" for value in item['leave_bed_times']])

def load_template_config(config):
    template_path = getattr(config, "LEAVE_BED_TEMPLATE_CONFIG", None)
    if not template_path or not os.path.exists(template_path):
        raise FileNotFoundError(f"找不到离床分析模板配置: {template_path}")
    with open(template_path, "r", encoding="utf-8") as f:
        return json.load(f)

def build_workbook_from_template_config(config, template_config, device_map):
    sheet_names = template_config.get("sheets") or ["Sheet1"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_names[0]
    for sheet_name in sheet_names[1:]:
        wb.create_sheet(sheet_name)

    active_name = template_config.get("active_sheet", sheet_names[0])
    ws = wb[active_name]

    for col, width in template_config.get("column_widths", {}).items():
        ws.column_dimensions[col].width = width

    default_align = Alignment(vertical="center", horizontal="left")
    data_align = Alignment(wrap_text=True, vertical="center", horizontal="left")
    thin = Side(style="thin")
    data_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    date_label_cell = template_config.get("date_label_cell", "A1")
    date_cell = template_config.get("date_cell", "C1")
    ws[date_label_cell] = "日期"
    ws[date_label_cell].alignment = default_align
    ws[date_cell] = config.file_date_obj
    ws[date_cell].number_format = template_config.get("date_number_format", "mm-dd-yy")
    ws[date_cell].alignment = default_align

    header_row = int(template_config.get("header_row", 2))
    for col, header in template_config.get("headers", {}).items():
        cell = ws[f"{col}{header_row}"]
        cell.value = header
        cell.alignment = default_align

    data_start_row = int(template_config.get("data_start_row", 3))
    data_row_height = template_config.get("data_row_height", 120)
    default_sleep_period = template_config.get("default_sleep_period", "入睡\n清醒")
    columns = template_config.get("columns", {})

    real_devices = [(device_id, info) for device_id, info in device_map.items() if isinstance(info, dict)]
    for offset, (device_id, info) in enumerate(real_devices):
        row = data_start_row + offset
        ws.row_dimensions[row].height = data_row_height
        ws[f"{columns.get('name', 'A')}{row}"] = info.get("name", "")
        ws[f"{columns.get('device_id', 'B')}{row}"] = device_id
        ws[f"{columns.get('floor', 'C')}{row}"] = info.get("floor", "")
        ws[f"{columns.get('sleep_period', 'D')}{row}"] = default_sleep_period

        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.alignment = data_align
            cell.border = data_border
        ws[f"{columns.get('sleep_period', 'D')}{row}"].number_format = "@"

    return wb, ws

def run(config, accurate_leave_data):
    logging.info("=== 开始生成 Excel 离床分析报告 ===")
    
    save_excel_path = os.path.join(config.BASE_DATA_PATH, f'合肥院离床数据分析{config.DATE_STR}晚.xlsx')
    if os.path.exists(save_excel_path):
        try:
            with open(save_excel_path, "r+b"): pass
        except PermissionError:
            logging.error(f"❌ 请先关闭Excel文件！{save_excel_path}")
            return

    device_map = get_device_mapping(config.DEVICE_ID_PATH, region=config.LOCATION_CONFIG.get("name"))
    template_config = load_template_config(config)
    
    csv_files = [f for f in os.listdir(config.WARN_DIR) if f.endswith('.csv')]
    alert_csv_path = os.path.join(config.WARN_DIR, csv_files[0]) if csv_files else None

    alert_data = defaultdict(list)
    if alert_csv_path:
        with open(alert_csv_path, 'r', encoding='utf-8-sig') as f:
            for row in csv.DictReader(f):
                if row['告警名称'].strip() == '离床告警':
                    try:
                        alert_data[clean_name(row['姓名'])].append(datetime.strptime(row['告警时间'].strip(), "%Y-%m-%d %H:%M:%S"))
                    except Exception: pass
        for name in alert_data: alert_data[name].sort()

    sleep_txt_path = os.path.join(config.REPORT_DIR, "睡眠报告.txt")
    sleep_data, unprocessed_diagnostics = {}, {}
    pattern = re.compile(r'([^\|]+)\s*\|\s*([^\|]+)\s*\|\s*入睡\s*([^\|]+)\s*\|\s*清醒\s*([^\|]+)\s*\|\s*离床时间：(.*)')
    
    if os.path.exists(sleep_txt_path):
        is_diagnostic = False
        with open(sleep_txt_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line: continue
                if "初步系统诊断如下" in line: is_diagnostic = True; continue
                
                if is_diagnostic and line.startswith("姓名："):
                    parts = line.split('|')
                    if len(parts) >= 3:
                        n = clean_name(parts[0].replace('姓名：', ''))
                        d = parts[2].replace('诊断结果：', '').strip()
                        if "无人在床" in d: unprocessed_diagnostics[n] = "无人在床"
                        elif "设备离线" in d: unprocessed_diagnostics[n] = "设备离线"
                        elif "需人工排查" in d: unprocessed_diagnostics[n] = "需要分析"
                    continue
                    
                match = pattern.match(line)
                if match:
                    name = clean_name(match.group(1))
                    leave_bed = match.group(5).strip().replace('、', '\n')
                    _append_sleep_record(
                        sleep_data,
                        name,
                        match.group(2).strip(),
                        match.group(3).strip(),
                        match.group(4).strip(),
                        leave_bed,
                    )

    _finalize_sleep_data(sleep_data)

    wb, ws = build_workbook_from_template_config(config, template_config, device_map)
    align = Alignment(wrap_text=True, vertical='center', horizontal='left')
    offline_fill = template_config.get("offline_fill", "00B0F0")
    blue_fill = PatternFill(start_color=offline_fill, end_color=offline_fill, fill_type="solid")
    columns = template_config.get("columns", {})
    data_start_row = int(template_config.get("data_start_row", 3))
    image_layout = template_config.get("image_layout", {})
    
    count_offline = count_empty = 0
    ref_year = config.file_date_obj.year

    for row in range(data_start_row, ws.max_row + 1):
        name_cell = ws[f"{columns.get('name', 'A')}{row}"].value or ""
        device_id = str(ws[f"{columns.get('device_id', 'B')}{row}"].value).strip() if ws[f"{columns.get('device_id', 'B')}{row}"].value else ""
        if not name_cell and not device_id:
            continue

        pure_name = clean_name(name_cell)
        
        alert_dts = alert_data.get(pure_name, [])
        ws[f"{columns.get('alerts', 'F')}{row}"] = '\n'.join([dt.strftime("%Y-%m-%d %H:%M:%S") for dt in alert_dts])

        if pure_name in sleep_data:
            ws[f"{columns.get('floor', 'C')}{row}"] = sleep_data[pure_name]['floor']
            ws[f"{columns.get('sleep_period', 'D')}{row}"] = sleep_data[pure_name]['入睡清醒']
            
            if device_id in accurate_leave_data and accurate_leave_data[device_id]:
                accurate_events = accurate_leave_data[device_id]
                ws[f"{columns.get('sleep_leave_bed', 'E')}{row}"] = '\n'.join([f"{ev['start_time']} 离床{ev['duration']}分钟" for ev in accurate_events])
            else:
                accurate_events = []
                ws[f"{columns.get('sleep_leave_bed', 'E')}{row}"] = sleep_data[pure_name]['e_离床时间']
                if sleep_data[pure_name]['离床时间']:
                    for ft in sleep_data[pure_name]['离床时间'].split('\n'):
                        if ft.strip(): accurate_events.append({'start_time': ft.strip(), 'duration': 0})

            ws[f"{columns.get('analysis', 'H')}{row}"] = '\n'.join(generate_diagnostics(sleep_data[pure_name], accurate_events, alert_dts, ref_year))
        else:
            ws[f"{columns.get('sleep_period', 'D')}{row}"] = ws[f"{columns.get('sleep_leave_bed', 'E')}{row}"] = ""
            if pure_name in unprocessed_diagnostics:
                diag_status = unprocessed_diagnostics[pure_name]
                ws[f"{columns.get('analysis', 'H')}{row}"] = diag_status
                if diag_status in ["无人在床", "设备离线"]:
                    count_empty += (diag_status == "无人在床")
                    count_offline += (diag_status == "设备离线")
                    for col in 'ABCDEFGH': ws[f'{col}{row}'].fill = blue_fill

        if device_id:
            device_folder = os.path.join(config.PLOT_DIR, device_id)
            # 兼容读取 body_status 和 24h_hr_summary
            img_paths = [os.path.join(r, f) for r, _, fs in os.walk(device_folder) for f in fs if f.endswith('.png') and ('body_status' in f or '24h_hr_summary' in f)]
            
            # 使用 sorted 排序，确保 00_ 开头的图排在第一张
            for i, p in enumerate(sorted(img_paths)):
                try:
                    img = Image(p)
                    img.width = image_layout.get("width_px", 100)
                    img.height = image_layout.get("height_px", 60)
                    images_per_row = image_layout.get("images_per_row", 3)
                    target_col = image_layout.get("target_column_index_zero_based", 6)
                    col_offset = image_layout.get("left_padding_px", 10) + (i % images_per_row) * image_layout.get("x_step_px", 105)
                    row_offset = image_layout.get("top_padding_px", 10) + (i // images_per_row) * image_layout.get("y_step_px", 65)
                    marker = AnchorMarker(col=target_col, colOff=int(col_offset*9525), row=row-1, rowOff=int(row_offset*9525))
                    img.anchor = OneCellAnchor(_from=marker, ext=XDRPositiveSize2D(cx=img.width*9525, cy=img.height*9525))
                    ws.add_image(img)
                except Exception: pass

        for col in [
            columns.get('sleep_period', 'D'),
            columns.get('sleep_leave_bed', 'E'),
            columns.get('alerts', 'F'),
            columns.get('analysis', 'H'),
        ]:
            ws[f'{col}{row}'].alignment = align

    wb.save(save_excel_path)
    logging.info(f"📊 【今日异常统计】 -> 设备离线: {count_offline} 台 | 无人在床: {count_empty} 台")
    logging.info(f"✅ Excel 报告生成在：{save_excel_path}")
