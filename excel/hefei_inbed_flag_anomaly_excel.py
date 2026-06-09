import argparse
import csv
import glob
import os
from datetime import datetime, time, timedelta

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font


DEFAULT_HEFEI_OUTPUT_ROOT = r"C:\Users\Lenovo\Desktop\data\合肥\output"
DEFAULT_REPORT_PATH = os.path.join(os.path.dirname(__file__), "合肥离床标签异常_睡眠时段汇总.xlsx")
SUMMARY_PATTERN = os.path.join("*", "inbed_flag_anomaly", "inbed_flag_anomaly_overview.csv")
SLEEP_START = time(21, 0, 0)
SLEEP_END = time(6, 0, 0)
MIN_DURATION_MINUTES = 5
IMAGE_WIDTH_PX = 620
IMAGE_HEIGHT_PX = 230


def _parse_dt(value):
    return datetime.strptime(value.strip(), "%Y-%m-%d %H:%M:%S")


def _sleep_windows_for_segment(start_dt, end_dt):
    dates = {
        (start_dt - timedelta(days=1)).date(),
        start_dt.date(),
        end_dt.date(),
        (end_dt + timedelta(days=1)).date(),
    }
    for day in sorted(dates):
        sleep_start = datetime.combine(day, SLEEP_START)
        sleep_end = datetime.combine(day + timedelta(days=1), SLEEP_END)
        yield sleep_start, sleep_end


def _overlaps_sleep_time(start_dt, end_dt):
    return any(start_dt <= sleep_end and end_dt >= sleep_start for sleep_start, sleep_end in _sleep_windows_for_segment(start_dt, end_dt))


def _duration_minutes(row, start_dt, end_dt):
    try:
        return float(row.get("持续分钟", ""))
    except ValueError:
        return (end_dt - start_dt).total_seconds() / 60


def _read_matching_rows(hefei_output_root):
    rows = []
    hefei_output_root = os.path.abspath(hefei_output_root)
    pattern = os.path.join(hefei_output_root, SUMMARY_PATTERN)
    for summary_path in sorted(glob.glob(pattern)):
        summary_path = os.path.abspath(summary_path)
        if not os.path.commonpath([hefei_output_root, summary_path]) == hefei_output_root:
            continue
        summary_dir = os.path.dirname(summary_path)
        with open(summary_path, "r", encoding="utf-8-sig", newline="") as f:
            for row in csv.DictReader(f):
                try:
                    start_dt = _parse_dt(row["开始时间"])
                    end_dt = _parse_dt(row["结束时间"])
                    duration_min = _duration_minutes(row, start_dt, end_dt)
                except Exception:
                    continue

                if duration_min <= MIN_DURATION_MINUTES:
                    continue
                if not _overlaps_sleep_time(start_dt, end_dt):
                    continue

                image_path = os.path.abspath(os.path.join(summary_dir, row.get("图片文件", "")))
                if not os.path.commonpath([hefei_output_root, image_path]) == hefei_output_root:
                    continue
                if not os.path.exists(image_path):
                    continue

                rows.append({
                    "device_id": row.get("设备号", ""),
                    "start_dt": start_dt,
                    "end_dt": end_dt,
                    "duration_min": duration_min,
                    "image_path": image_path,
                })
    rows.sort(key=lambda item: (item["start_dt"], item["device_id"]))
    return rows


def _format_period(row):
    return (
        f"{row['start_dt'].strftime('%Y-%m-%d %H:%M:%S')}\n"
        f"至 {row['end_dt'].strftime('%Y-%m-%d %H:%M:%S')}\n"
        f"持续 {row['duration_min']:.2f} 分钟"
    )


def generate_report(hefei_output_root=DEFAULT_HEFEI_OUTPUT_ROOT, report_path=DEFAULT_REPORT_PATH):
    rows = _read_matching_rows(hefei_output_root)
    os.makedirs(os.path.dirname(report_path), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "合肥睡眠时段异常"

    headers = ["设备号", "异常时间段", "数值状态图"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = min(255, IMAGE_WIDTH_PX / 7.5 + 4)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:C1"

    text_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    image_row_height = min(409, IMAGE_HEIGHT_PX * 0.75 + 12)

    for row_idx, row in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=row["device_id"]).alignment = text_alignment
        ws.cell(row=row_idx, column=2, value=_format_period(row)).alignment = text_alignment
        ws.row_dimensions[row_idx].height = image_row_height

        image = ExcelImage(row["image_path"])
        image.width = IMAGE_WIDTH_PX
        image.height = IMAGE_HEIGHT_PX
        ws.add_image(image, f"C{row_idx}")

    wb.save(report_path)
    return report_path, len(rows)


def main():
    parser = argparse.ArgumentParser(description="聚合合肥睡眠时段 inbed_flag 异常图到 Excel")
    parser.add_argument("--hefei-output-root", default=DEFAULT_HEFEI_OUTPUT_ROOT, help="合肥 output 根目录")
    parser.add_argument("--report-path", default=DEFAULT_REPORT_PATH, help="输出 Excel 路径")
    args = parser.parse_args()

    report_path, row_count = generate_report(args.hefei_output_root, args.report_path)
    print(f"生成完成: {report_path}")
    print(f"写入条目: {row_count}")


if __name__ == "__main__":
    main()
