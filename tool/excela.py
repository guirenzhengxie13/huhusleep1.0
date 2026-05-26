import os
import glob
import io
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from PIL import Image as PILImage

def generate_excel_report():
    # 1. 路径配置
    output_dir = r"C:\Users\Lenovo\Desktop\datatest\画图\output"
    report_path = os.path.join(output_dir, "设备体征图表大巡检报告_原图高清自适应版.xlsx")

    # 创建新的 Excel 工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "图表汇总"

    # 设置表头样式
    ws['A1'] = "设备号"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.column_dimensions['A'].width = 25

    # ================= 🎯 排版与偏移配置区 =================
    OFFSET_X_PX = 15            # 图片向右偏移的距离（像素）- 左边距
    OFFSET_Y_PX = 15            # 图片向下偏移的距离（像素）- 上边距
    EXTRA_MARGIN_BOTTOM = 15    # 底部额外留白
    EXTRA_MARGIN_RIGHT = 15     # 右侧额外留白
    
    # 手动微调系数：如果你觉得算出来的格子还不够大，可以调大这两个值
    ROW_HEIGHT_ADJUST = 0       # 额外增加的行高（磅）
    COL_WIDTH_ADJUST = 2        # 额外增加的列宽（字符）
    # =======================================================

    print("🔍 开始扫描设备文件夹...")
    
    folder_names = [f for f in os.listdir(output_dir) if os.path.isdir(os.path.join(output_dir, f))]
    
    if not folder_names:
        print("⚠️ 未找到任何设备文件夹！请检查路径。")
        return

    row_idx = 2 
    max_col_used = 1 

    for folder_name in folder_names:
        folder_path = os.path.join(output_dir, folder_name)
        
        cell = ws.cell(row=row_idx, column=1, value=folder_name)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        image_files = glob.glob(os.path.join(folder_path, "*.png"))
        image_files.sort()
        
        if not image_files:
            continue
            
        print(f"✅ 正在处理设备: {folder_name}，发现 {len(image_files)} 张图表")

        for i, img_path in enumerate(image_files):
            col_idx = i + 2 
            col_letter = get_column_letter(col_idx)
            
            if col_idx > max_col_used:
                max_col_used = col_idx

            # 2. 读取原图尺寸并创建带边距的画布
            with PILImage.open(img_path) as original_img:
                orig_w, orig_h = original_img.size
                
                # 画布总大小 = 原图大小 + 偏移量 + 留白
                canvas_w = orig_w + OFFSET_X_PX + EXTRA_MARGIN_RIGHT
                canvas_h = orig_h + OFFSET_Y_PX + EXTRA_MARGIN_BOTTOM
                
                canvas = PILImage.new("RGB", (canvas_w, canvas_h), "white")
                canvas.paste(original_img, (OFFSET_X_PX, OFFSET_Y_PX))
                
                img_byte_arr = io.BytesIO()
                canvas.save(img_byte_arr, format='PNG')
                img_byte_arr.seek(0)

            # 3. 插入图片
            img = ExcelImage(img_byte_arr)
            cell_ref = f"{col_letter}{row_idx}"
            ws.add_image(img, cell_ref)

            # 4. ⭐️ 核心：带极限保护的行列宽高调整
            # 换算单位 (1 像素 ≈ 0.75 磅, 1 字符宽度 ≈ 7.5 像素)
            target_height_pt = (canvas_h * 0.75) + ROW_HEIGHT_ADJUST
            target_width_char = (canvas_w / 7.5) + COL_WIDTH_ADJUST

            # Excel 最大行高是 409 磅，最大列宽是 255 字符，必须加 min() 保护，否则会报错损坏文件
            final_height = min(target_height_pt, 409)
            final_width = min(target_width_char, 255)

            ws.row_dimensions[row_idx].height = final_height
            ws.column_dimensions[col_letter].width = final_width
            
            # 如果原图实在是太大了超过了 Excel 极限，给个友好提示
            if target_height_pt > 409:
                print(f"   ⚠️ 提示: 图片 {os.path.basename(img_path)} 的高度超出了 Excel 单个单元格的极限(409磅)，单元格已尽力撑到最大。")

        row_idx += 1

    # 统一补充表头
    for col in range(2, max_col_used + 1):
        ws.cell(row=1, column=col, value=f"监测数据图 {col-1}").font = Font(bold=True)
        ws.cell(row=1, column=col).alignment = Alignment(horizontal='center')

    wb.save(report_path)
    print(f"\n🎉 大功告成！完美适应原图宽高的报告已生成：\n{report_path}")

if __name__ == "__main__":
    generate_excel_report()
