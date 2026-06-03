import csv
import json
import shutil
import sys
from ast import literal_eval
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side


PROJECT_ROOT = Path(__file__).resolve().parents[1]
WEEKREPORT_ROOT = Path(__file__).resolve().parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from pipeline.device_matcher import (  # noqa: E402
    detect_group,
    extract_device_ids_from_csv,
)
from utils import get_device_roster  # noqa: E402


BASE_DATA_DIR = Path.home() / "Desktop" / "data" / "失能周报数据"
LOCAL_HISTORY_DATA_DIR = WEEKREPORT_ROOT / "失能周报数据"
CONFIG_FILE_PATH = WEEKREPORT_ROOT / "config.json"
LEGACY_CONFIG_FILE_PATH = Path.home() / "Desktop" / "代码" / "失能周报" / "config.json"
VERIFICATION_STORE_PATH = WEEKREPORT_ROOT / "verification_records.csv"
DEVICE_ROSTER_PATH = PROJECT_ROOT / "assets" / "full_device_roster.csv"
EXTRA_VERIFICATION_SOURCE_NAMES = {"失能老人识别活力老人情况核实.xlsx"}

now = datetime.now()
DATE_STR = f"{now.month}{now.strftime('%d')}"

RAW_IMPORT_DIR = Path.home() / "Downloads"
INPUT_DIR = BASE_DATA_DIR / DATE_STR
OUTPUT_DIR = BASE_DATA_DIR / f"{DATE_STR}output"
WEEKREPORT_REQUIRED_COLUMNS = {"deviceName", "eftv_times", "unabledetect_result", "report_date_week"}
VERIFICATION_COLUMNS = [
    "院区",
    "deviceName",
    "老人",
    "养老院失能评估",
    "设备失能评估",
    "在床时间",
    "平均离床次数",
    "协助离床占比",
    "平时离床情况",
    "每日离床次数范围",
    "主要离床方式",
    "设备评估合理性",
    "来源文件",
    "更新时间",
]
THIN_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)

COLUMN_WIDTHS = {
    "A": 4.75,
    "B": 8.75,
    "C": 17.875,
    "D": 12.0,
    "E": 16.0,
    "F": 12.875,
    "G": 48.25,
    "H": 24.875,
    "I": 24.125,
    "J": 24.875,
    "K": 30.375,
    "L": 48.25,
    "M": 56.0,
    "N": 24.875,
    "O": 52.625,
    "P": 53.75,
    "Q": 108.25,
    "R": 118.25,
    "S": 9.875,
    "T": 26.0,
}
ROW_HEIGHTS = {"1": 30.0}

RESULT_MAPPING = {
    0: "活力老人",
    1: "半失能老人",
    2: "重度失能老人",
}
PREDICTED_LEVELS = ["活力老人", "半失能老人", "重度失能老人"]
CARE_LEVEL_ORDER = ["活力老人", "轻度失能老人", "中度失能老人", "重度失能老人", "全失能老人"]
CARE_LEVEL_TO_STATS_LABEL = {
    "自理": "活力老人",
    "活力老人": "活力老人",
    "轻度失能": "轻度失能老人",
    "轻度失能老人": "轻度失能老人",
    "中度失能": "中度失能老人",
    "中度失能老人": "中度失能老人",
    "重度失能": "重度失能老人",
    "重度失能老人": "重度失能老人",
    "全失能": "全失能老人",
    "全失能老人": "全失能老人",
}
MISJUDGE_DISABLED_LEVELS = {"中度失能", "中度失能老人", "重度失能", "重度失能老人", "全失能", "全失能老人"}


def load_config_data():
    config_path = CONFIG_FILE_PATH if CONFIG_FILE_PATH.exists() else LEGACY_CONFIG_FILE_PATH
    if not config_path.exists():
        raise FileNotFoundError(f"找不到周报配置文件: {CONFIG_FILE_PATH}")

    with config_path.open("r", encoding="utf-8") as f:
        return json.load(f)


def normalize_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    return str(value).strip()


def load_verification_records():
    if not VERIFICATION_STORE_PATH.exists():
        return {}

    df = pd.read_csv(VERIFICATION_STORE_PATH, encoding="utf-8-sig", dtype=str).fillna("")
    records = {}
    for record in df.to_dict("records"):
        device_id = normalize_text(record.get("deviceName"))
        if device_id:
            records[device_id] = {column: normalize_text(record.get(column)) for column in VERIFICATION_COLUMNS}
    return records


def save_verification_records(records):
    VERIFICATION_STORE_PATH.parent.mkdir(parents=True, exist_ok=True)
    rows = []
    for device_id in sorted(records):
        row = {column: normalize_text(records[device_id].get(column)) for column in VERIFICATION_COLUMNS}
        row["deviceName"] = device_id
        rows.append(row)

    pd.DataFrame(rows, columns=VERIFICATION_COLUMNS).to_csv(
        VERIFICATION_STORE_PATH,
        index=False,
        encoding="utf-8-sig",
    )


def refresh_verification_store_from_outputs(config_data):
    records = load_verification_records()
    found_count = 0

    output_files = []
    for history_dir in (LOCAL_HISTORY_DATA_DIR, BASE_DATA_DIR):
        if history_dir.exists():
            output_files.extend(history_dir.glob("*output/*.xlsx"))

    desktop_dir = Path.home() / "Desktop"
    if desktop_dir.exists():
        for workbook_path in desktop_dir.rglob("*.xlsx"):
            if workbook_path.name in EXTRA_VERIFICATION_SOURCE_NAMES:
                output_files.append(workbook_path)

    output_files = sorted(set(output_files), key=lambda path: path.stat().st_mtime)
    for workbook_path in output_files:
        if workbook_path.name.startswith("~$"):
            continue
        for record in extract_verification_records_from_workbook(workbook_path, config_data):
            device_id = record["deviceName"]
            records[device_id] = record
            found_count += 1

    save_verification_records(records)
    print(f"核实映射已更新: 当前保存 {len(records)} 条，本次扫描吸收 {found_count} 条。")
    return records


def extract_verification_records_from_workbook(workbook_path, config_data):
    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception as exc:
        print(f"警告: 跳过核实数据扫描 {workbook_path.name}: {exc}")
        return []

    worksheets = find_verification_worksheets(workbook)
    if not worksheets:
        return []

    device_to_region = build_device_to_region(config_data)
    device_by_name = build_device_by_name(config_data)
    updated_at = datetime.fromtimestamp(workbook_path.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S")

    records = []
    for worksheet, bottom_header_row in worksheets:
        top_device_map = build_misjudged_top_device_map(worksheet, bottom_header_row)
        blank_streak = 0

        for row_idx in range(bottom_header_row + 2, worksheet.max_row + 1):
            seq = normalize_text(worksheet.cell(row=row_idx, column=1).value)
            elder_name = normalize_text(worksheet.cell(row=row_idx, column=2).value)
            if not seq and not elder_name:
                blank_streak += 1
                if blank_streak >= 5:
                    break
                continue
            blank_streak = 0

            device_info = top_device_map.get(seq)
            if not device_info and elder_name:
                device_info = device_by_name.get(elder_name)
            if not device_info:
                continue

            manual_values = {
                "在床时间": normalize_text(worksheet.cell(row=row_idx, column=5).value),
                "平均离床次数": normalize_text(worksheet.cell(row=row_idx, column=6).value),
                "协助离床占比": normalize_text(worksheet.cell(row=row_idx, column=7).value),
                "平时离床情况": normalize_text(worksheet.cell(row=row_idx, column=8).value),
                "每日离床次数范围": normalize_text(worksheet.cell(row=row_idx, column=9).value),
                "主要离床方式": normalize_text(worksheet.cell(row=row_idx, column=10).value),
                "设备评估合理性": normalize_text(worksheet.cell(row=row_idx, column=11).value),
            }
            if not any(
                manual_values[column]
                for column in ["平时离床情况", "每日离床次数范围", "主要离床方式", "设备评估合理性"]
            ):
                continue

            device_id = device_info["deviceName"]
            records.append(
                {
                    "院区": device_to_region.get(device_id, ""),
                    "deviceName": device_id,
                    "老人": elder_name or device_info["老人"],
                    "养老院失能评估": normalize_text(worksheet.cell(row=row_idx, column=3).value) or device_info["养老院失能评估"],
                    "设备失能评估": normalize_text(worksheet.cell(row=row_idx, column=4).value) or device_info["设备失能评估"],
                    **manual_values,
                    "来源文件": str(workbook_path),
                    "更新时间": updated_at,
                }
            )
    return records


def find_verification_worksheets(workbook):
    worksheets = []
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        header_row = find_verification_header_row(worksheet)
        if header_row:
            worksheets.append((worksheet, header_row))
    return worksheets


def find_verification_header_row(worksheet):
    for row_idx in range(1, worksheet.max_row + 1):
        first = normalize_text(worksheet.cell(row=row_idx, column=1).value)
        second = normalize_text(worksheet.cell(row=row_idx, column=2).value)
        third = normalize_text(worksheet.cell(row=row_idx, column=3).value)
        if first == "序号" and second == "老人" and third == "养老院失能评估":
            return row_idx
    return None


def build_misjudged_top_device_map(worksheet, bottom_header_row):
    mapping = {}
    for row_idx in range(2, bottom_header_row):
        seq = normalize_text(worksheet.cell(row=row_idx, column=1).value)
        device_id = normalize_text(worksheet.cell(row=row_idx, column=3).value)
        if not seq or not device_id:
            continue

        mapping[seq] = {
            "老人": normalize_text(worksheet.cell(row=row_idx, column=2).value),
            "deviceName": device_id,
            "养老院失能评估": normalize_text(worksheet.cell(row=row_idx, column=4).value),
            "设备失能评估": normalize_text(worksheet.cell(row=row_idx, column=5).value),
        }
    return mapping


def build_device_to_region(config_data):
    return {
        normalize_text(row["设备号"]): normalize_text(row["院区"])
        for row in get_device_roster(DEVICE_ROSTER_PATH)
        if normalize_text(row["设备号"]) and normalize_text(row["院区"])
    }


def build_care_level_by_device(config_data):
    care_level_by_device = {}
    for rows in config_data.get("regions", {}).values():
        for row in rows:
            if len(row) >= 4:
                device_id = normalize_text(row[2])
                care_level = normalize_text(row[3])
                if device_id and care_level:
                    care_level_by_device[device_id] = care_level
    return care_level_by_device


def build_regions_from_roster(config_data):
    care_level_by_device = build_care_level_by_device(config_data)
    regions = defaultdict(list)
    counters = defaultdict(int)

    for row in get_device_roster(DEVICE_ROSTER_PATH):
        region_name = normalize_text(row["院区"])
        device_id = normalize_text(row["设备号"])
        if not region_name or not device_id:
            continue

        counters[region_name] += 1
        regions[region_name].append([
            str(counters[region_name]),
            normalize_text(row["老人姓名"]),
            device_id,
            normalize_text(row["失能等级"]) or care_level_by_device.get(device_id, ""),
        ])

    return dict(regions)


def build_device_by_name(config_data):
    device_by_name = {}
    care_level_by_device = build_care_level_by_device(config_data)
    for row in get_device_roster(DEVICE_ROSTER_PATH):
        elder_name = normalize_text(row["老人姓名"])
        device_id = normalize_text(row["设备号"])
        if elder_name and device_id:
            device_by_name[elder_name] = {
                "院区": normalize_text(row["院区"]),
                "deviceName": device_id,
                "老人": elder_name,
                "养老院失能评估": normalize_text(row["失能等级"]) or care_level_by_device.get(device_id, ""),
                "设备失能评估": "",
            }
    return device_by_name


def read_csv_safely(csv_path):
    for encoding in ("utf-8-sig", "utf-8", "gbk"):
        try:
            return pd.read_csv(csv_path, encoding=encoding)
        except UnicodeDecodeError:
            continue
    return pd.read_csv(csv_path)


def is_weekreport_csv(csv_path):
    for encoding in ("utf-8-sig", "utf-8", "gbk"):
        try:
            with open(csv_path, "r", encoding=encoding, errors="ignore", newline="") as f:
                for row in csv.reader(f):
                    normalized = {str(cell).strip().lstrip("\ufeff") for cell in row}
                    if not any(normalized):
                        continue
                    return WEEKREPORT_REQUIRED_COLUMNS.issubset(normalized)
        except UnicodeDecodeError:
            continue
    return False


def archive_downloaded_csvs(config_data, replace_existing=True):
    device_to_region = build_device_to_region(config_data)
    if not device_to_region:
        print("警告: 设备总表中没有可用于匹配的设备号，跳过自动归档。")
        return []

    INPUT_DIR.mkdir(parents=True, exist_ok=True)
    archived = []
    skipped = []

    for csv_path in sorted(RAW_IMPORT_DIR.glob("*.csv")):
        if not is_weekreport_csv(csv_path):
            continue

        device_ids = extract_device_ids_from_csv(str(csv_path))
        region_name, counts = detect_group(device_ids, device_to_region)
        if not region_name:
            skipped.append((csv_path, "未匹配到设备总表中的设备号"))
            continue

        if len(counts) > 1:
            archived.extend(_split_mixed_region_csv(csv_path, device_to_region, replace_existing))
            continue

        target_path = INPUT_DIR / f"{region_name}.csv"
        if target_path.exists() and not replace_existing:
            skipped.append((csv_path, f"目标已存在: {target_path}"))
            continue

        shutil.copy2(str(csv_path), str(target_path))
        archived.append((csv_path, target_path, region_name, len(device_ids)))

    for csv_path, reason in skipped:
        print(f"警告: 跳过 {csv_path.name}: {reason}")
    for source, target, region_name, count in archived:
        print(f"完成: 已识别 [{region_name}] {source.name}，设备记录 {count} 条，已复制到 -> {target}")
    return archived


def _split_mixed_region_csv(csv_path, device_to_region, replace_existing):
    df = read_csv_safely(csv_path)

    if "deviceName" not in df.columns:
        print(f"警告: 跳过混合文件 {csv_path.name}: 缺少 deviceName 列")
        return []

    df["_region"] = df["deviceName"].astype(str).map(device_to_region)
    archived = []
    source_dir = INPUT_DIR / "source_files"
    source_dir.mkdir(parents=True, exist_ok=True)

    for region_name, region_df in df.dropna(subset=["_region"]).groupby("_region"):
        target_path = INPUT_DIR / f"{region_name}.csv"
        if target_path.exists() and not replace_existing:
            print(f"警告: 跳过 [{region_name}] 拆分输出，目标已存在: {target_path}")
            continue
        region_df.drop(columns=["_region"]).to_csv(target_path, index=False, encoding="utf-8-sig")
        archived.append((csv_path, target_path, region_name, len(region_df)))

    archived_source = source_dir / csv_path.name
    shutil.copy2(str(csv_path), str(archived_source))
    return archived


def build_region_dataframe(device_list, columns, input_csv_path):
    df_names = pd.DataFrame(device_list, columns=columns)
    df_sql = read_csv_safely(input_csv_path)
    if "latest_log_time" in df_sql.columns:
        df_sql = df_sql.drop(columns=["latest_log_time"])

    df_sql.replace(" ", np.nan, inplace=True)
    df_final = pd.merge(df_names, df_sql, on="deviceName", how="left")

    if "unabledetect_result" in df_final.columns:
        numeric_result = pd.to_numeric(df_final["unabledetect_result"], errors="coerce")
        df_final["unabledetect_result_mapped"] = numeric_result.map(RESULT_MAPPING)
        df_final = move_column_after(df_final, "unabledetect_result_mapped", "护理级别")

    return df_final


def move_column_after(df, column_name, after_column):
    if column_name not in df.columns or after_column not in df.columns:
        return df

    cols = list(df.columns)
    cols.remove(column_name)
    insert_at = cols.index(after_column) + 1
    cols.insert(insert_at, column_name)
    return df[cols]


def filter_valid_devices(df):
    if "eftv_times" not in df.columns:
        return df.iloc[0:0].copy()

    valid_days = pd.to_numeric(df["eftv_times"], errors="coerce")
    return df[valid_days > 5].copy()


def extract_misjudged_data(valid_df):
    if "护理级别" not in valid_df.columns or "unabledetect_result_mapped" not in valid_df.columns:
        return valid_df.iloc[0:0].copy()

    care_level = valid_df["护理级别"].astype(str).str.strip()
    predicted = valid_df["unabledetect_result_mapped"].astype(str).str.strip()

    self_judged_disabled = care_level.eq("自理") & predicted.isin(["半失能老人", "重度失能老人"])
    disabled_judged_self = care_level.isin(MISJUDGE_DISABLED_LEVELS) & predicted.eq("活力老人")
    return valid_df[self_judged_disabled | disabled_judged_self].copy()


def build_stats_table(valid_df, region_name):
    if "护理级别" not in valid_df.columns or "unabledetect_result_mapped" not in valid_df.columns:
        return pd.DataFrame()

    actual = valid_df["护理级别"].astype(str).str.strip().map(CARE_LEVEL_TO_STATS_LABEL)
    predicted = valid_df["unabledetect_result_mapped"].astype(str).str.strip()
    stats_source = pd.DataFrame({"养老院评估": actual, "设备识别": predicted}).dropna()

    matrix = pd.crosstab(stats_source["养老院评估"], stats_source["设备识别"])
    matrix = matrix.reindex(index=CARE_LEVEL_ORDER, columns=PREDICTED_LEVELS, fill_value=0)
    matrix.insert(0, "", matrix.index)
    matrix.reset_index(drop=True, inplace=True)
    matrix.columns = [f"{region_name}院{DATE_STR}", "识别为活力老人", "识别为半失能老人", "识别为重度失能老人"]

    active_total = int(matrix.loc[matrix.iloc[:, 0].eq("活力老人"), matrix.columns[1:]].sum(axis=1).sum())
    active_correct = int(matrix.loc[matrix.iloc[:, 0].eq("活力老人"), "识别为活力老人"].sum())
    active_accuracy = active_correct / active_total if active_total else 0

    non_light_disabled = matrix.iloc[:, 0].isin(["中度失能老人", "重度失能老人", "全失能老人"])
    non_self = matrix.iloc[:, 0].isin(["轻度失能老人", "中度失能老人", "重度失能老人", "全失能老人"])
    false_self = int(matrix.loc[non_light_disabled, "识别为活力老人"].sum())
    non_self_total = int(matrix.loc[non_self, matrix.columns[1:]].sum(axis=1).sum())
    false_self_rate = false_self / non_self_total if non_self_total else 0

    summary_rows = pd.DataFrame(
        [
            ["", "", "", ""],
            [f"活力老人识别准确率：{active_accuracy:.2%}；", "", "", ""],
            [f"误报率（不包括轻度失能老人）：{false_self_rate:.1%}", "", "", ""],
        ],
        columns=matrix.columns,
    )
    return pd.concat([matrix, summary_rows], ignore_index=True)


def write_report_workbook(output_excel_path, report_df, misjudged_df, stats_df, region_name, verification_records):
    with pd.ExcelWriter(output_excel_path, engine="openpyxl") as writer:
        report_df.to_excel(writer, index=False, sheet_name="报表")
        misjudged_df.to_excel(writer, index=False, sheet_name="误判")
        stats_df.to_excel(writer, index=False, sheet_name="误报率")

        apply_report_sheet_style(writer.sheets["报表"])
        apply_misjudged_sheet_style(writer.sheets["误判"], misjudged_df)
        append_misjudged_analysis(writer.sheets["误判"], misjudged_df, verification_records)

        apply_stats_sheet_style(writer.sheets["误报率"])
        for worksheet in writer.sheets.values():
            expand_sheet(worksheet)

    print(
        f"完成: [{region_name}] 原始设备 {len(report_df)} 条，"
        f"误判 {len(misjudged_df)} 条，已保存: {output_excel_path}"
    )


def apply_report_sheet_style(ws):
    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    for row_num, height in ROW_HEIGHTS.items():
        ws.row_dimensions[int(row_num)].height = height

    center_alignment = Alignment(horizontal="center", vertical="center")
    for col_idx in range(1, min(7, ws.max_column) + 1):
        for col in ws.iter_cols(min_col=col_idx, max_col=col_idx):
            for cell in col:
                cell.alignment = center_alignment

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = center_alignment

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    apply_border_range(ws, 1, 1, ws.max_row, ws.max_column)


def apply_misjudged_sheet_style(ws, misjudged_df):
    widths = {
        "A": 5.375,
        "B": 9.375,
        "C": 18.25,
        "D": 12.875,
        "E": 32.875,
        "F": 12.875,
        "G": 48.25,
        "H": 24.875,
        "I": 24.125,
        "J": 22.875,
        "K": 30.375,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    center_alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=1, max_row=max(1, len(misjudged_df) + 1)):
        for cell in row:
            cell.alignment = center_alignment

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = center_alignment

    apply_border_range(ws, 1, 1, max(1, len(misjudged_df) + 1), ws.max_column)


def append_misjudged_analysis(ws, misjudged_df, verification_records):
    start_row = len(misjudged_df) + 6
    header_row = [
        "序号",
        "老人",
        "养老院失能评估",
        "设备失能评估",
        "设备数据分析",
        None,
        None,
        "老人情况核实",
        None,
        None,
        "设备评估合理性",
    ]
    subheader_row = [
        None,
        None,
        None,
        None,
        "在床时间",
        "平均离床次数",
        "协助离床占比",
        "平时离床情况",
        "每日离床次数范围",
        "主要离床方式",
        None,
    ]

    for col_idx, value in enumerate(header_row, start=1):
        ws.cell(row=start_row, column=col_idx, value=value)
    for col_idx, value in enumerate(subheader_row, start=1):
        ws.cell(row=start_row + 1, column=col_idx, value=value)

    merge_ranges = [
        f"A{start_row}:A{start_row + 1}",
        f"B{start_row}:B{start_row + 1}",
        f"C{start_row}:C{start_row + 1}",
        f"D{start_row}:D{start_row + 1}",
        f"E{start_row}:G{start_row}",
        f"H{start_row}:J{start_row}",
        f"K{start_row}:K{start_row + 1}",
    ]
    for merge_range in merge_ranges:
        ws.merge_cells(merge_range)

    for offset, (_, row) in enumerate(misjudged_df.iterrows(), start=2):
        target_row = start_row + offset
        analysis = build_misjudged_analysis_row(row, verification_records)
        for col_idx, value in enumerate(analysis, start=1):
            ws.cell(row=target_row, column=col_idx, value=value)

    center_alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=start_row, max_row=start_row + 1, min_col=1, max_col=11):
        for cell in row:
            cell.alignment = center_alignment
            cell.font = Font(bold=True)

    for row in ws.iter_rows(min_row=start_row + 2, max_row=start_row + len(misjudged_df) + 1, min_col=1, max_col=11):
        for cell in row:
            cell.alignment = center_alignment

    apply_border_range(ws, start_row, 1, start_row + len(misjudged_df) + 1, 11)


def build_misjudged_analysis_row(row, verification_records):
    inbed_hours = average_list_value(row.get("inbedtimes_week")) / 3600
    leave_counts = parse_number_list(row.get("num_leave_week"))
    help_counts = parse_number_list(row.get("helpsituations_week"))

    avg_leave = int(sum(leave_counts) / len(leave_counts)) if leave_counts else None
    help_ratio = sum(help_counts) / sum(leave_counts) if leave_counts and sum(leave_counts) else 0

    device_id = normalize_text(row.get("deviceName"))
    verification = verification_records.get(device_id, {})

    return [
        row.get("序号"),
        row.get("入住姓名"),
        row.get("护理级别"),
        row.get("unabledetect_result_mapped"),
        format_hour_bucket(inbed_hours),
        avg_leave,
        help_ratio,
        verification.get("平时离床情况") or None,
        verification.get("每日离床次数范围") or None,
        verification.get("主要离床方式") or None,
        verification.get("设备评估合理性") or None,
    ]


def parse_number_list(value):
    if pd.isna(value):
        return []
    if isinstance(value, (list, tuple)):
        raw_values = value
    else:
        text = str(value).strip()
        if not text:
            return []
        try:
            raw_values = literal_eval(text)
        except (ValueError, SyntaxError):
            return []

    numbers = []
    for item in raw_values:
        try:
            numbers.append(float(item))
        except (TypeError, ValueError):
            continue
    return numbers


def average_list_value(value):
    numbers = parse_number_list(value)
    return sum(numbers) / len(numbers) if numbers else 0


def format_hour_bucket(hours):
    if hours <= 0:
        return None
    start_hour = int(hours // 2 * 2)
    end_hour = start_hour + 2
    return f"{start_hour}-{end_hour}h"


def apply_stats_sheet_style(ws):
    widths = {"A": 22, "B": 18, "C": 20, "D": 20}
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    center_alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = center_alignment
            cell.border = THIN_BORDER

    for cell in ws[1]:
        cell.font = Font(bold=True)


def apply_border_range(ws, min_row, min_col, max_row, max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = THIN_BORDER


def expand_sheet(ws):
    for row_idx in list(ws.row_dimensions):
        ws.row_dimensions[row_idx].hidden = False

    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].hidden = False

    for col_letter in list(ws.column_dimensions):
        ws.column_dimensions[col_letter].hidden = False

    for col_idx in range(1, ws.max_column + 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].hidden = False


def process_region_report(region_name, device_list, columns, input_csv_path, output_excel_path, verification_records):
    print(f"开始处理 [{region_name}] 地区数据...")

    try:
        df_final = build_region_dataframe(device_list, columns, input_csv_path)
    except Exception as e:
        print(f"错误: 读取或合并 {region_name} CSV 失败: {e}")
        return

    valid_df = filter_valid_devices(df_final)
    misjudged_df = extract_misjudged_data(valid_df)
    stats_df = build_stats_table(valid_df, region_name)
    output_excel_path = available_output_path(output_excel_path)
    write_report_workbook(output_excel_path, df_final, misjudged_df, stats_df, region_name, verification_records)


def available_output_path(output_excel_path):
    if not output_excel_path.exists():
        return output_excel_path

    try:
        with open(output_excel_path, "a+b"):
            return output_excel_path
    except PermissionError:
        suffix = datetime.now().strftime("%H%M%S")
        fallback_path = output_excel_path.with_name(f"{output_excel_path.stem}_new_{suffix}{output_excel_path.suffix}")
        print(f"警告: 输出文件被占用，改写到: {fallback_path}")
        return fallback_path


def main():
    INPUT_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    config_data = load_config_data()
    verification_records = refresh_verification_store_from_outputs(config_data)
    archive_downloaded_csvs(config_data)

    columns = config_data.get("columns", ["序号", "入住姓名", "deviceName", "护理级别"])
    regions_data = build_regions_from_roster(config_data)
    if not regions_data:
        print("警告: 设备总表中没有找到可用的院区设备。")
        return

    for region_name, device_list in regions_data.items():
        input_csv = INPUT_DIR / f"{region_name}.csv"
        output_excel = OUTPUT_DIR / f"{DATE_STR}{region_name}周报.xlsx"

        if input_csv.exists():
            process_region_report(region_name, device_list, columns, input_csv, output_excel, verification_records)
        else:
            print(f"警告: 跳过 [{region_name}]: 未在 {INPUT_DIR} 中找到 {region_name}.csv")

    print("\n所有周报任务执行完毕。")


if __name__ == "__main__":
    main()
